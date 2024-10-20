import openpyxl
book = openpyxl.load_workbook("C:\\Users\\vener\\Documents\\pythondemo.xlsx")
sheet = book.active
Dict = {} #Empty dictionary


for i in range(1, sheet.max_row+1): # to get rows
    if(sheet.cell(row=i,column=1).value == "TestCase 2"):
        for j in range(2,sheet.max_column+1): # to get columns

            # Dict["firstname"]="Vikash"
            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

print(Dict)
