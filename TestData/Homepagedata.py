import openpyxl


class HomePageData:

    test_homePage_data = [{'firstname':'Vikash', 'lastname':'Dwivedi', 'gender':'Male'}, {'firstname':'Abhiranjan', 'lastname':'Singh', 'gender':'Male'}]

    @staticmethod
    def getExcelTestData(Test_case_name):
        Dict = {}  # Empty dictionary
        book = openpyxl.load_workbook("C:\\Users\\vener\\Documents\\pythondemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if (sheet.cell(row=i, column=1).value == Test_case_name):
                for j in range(2, sheet.max_column + 1):  # to get columns

                    # Dict["firstname"]="Vikash"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]

"""To call any method directly with classname without creating an objetct of the call 
simply make that method as static method see above to understand."""
# Self parameter required only with non-static method.